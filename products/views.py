# from rest_framework.generics import ListAPIView
# from rest_framework.filters import SearchFilter, OrderingFilter
# from .models import Product
# from .serializers import ProductSerializer
#
# class ProductListView(ListAPIView):
#     serializer_class = ProductSerializer
#     queryset = Product.objects.select_related('category').all()
#     filter_backends = [SearchFilter, OrderingFilter]
#     search_fields = ['name', 'category__name']
#     ordering_fields = ['price', 'name', 'stock']
#     ordering = ['name']
#
#     def get_queryset(self):
#         qs = super().get_queryset()
#         category = self.request.query_params.get('category')
#         if category:
#             qs = qs.filter(category__name__iexact=category)
#         return qs
#
#
#
# products/views.py
from io import BytesIO
from datetime import datetime, date, timedelta
from django.utils.timezone import make_aware, get_current_timezone
from django.db import transaction
from django.db.models import Sum, Count
from django.http import HttpResponse, FileResponse
from rest_framework.views import APIView
from rest_framework.generics import ListAPIView
from rest_framework.permissions import IsAdminUser
from rest_framework.response import Response
from rest_framework import status
from rest_framework.filters import SearchFilter, OrderingFilter
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from billing.models import BillingInvoice
from .models import Product
from .serializers import (
    ProductSerializer,
    ProductBulkUpdateSerializer,
)

# ---------- Existing products list ----------
class ProductListView(ListAPIView):
    serializer_class = ProductSerializer
    queryset = Product.objects.select_related('category').all()
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ['name', 'category__name']
    ordering_fields = ['price', 'name', 'stock']
    ordering = ['name']

    def get_queryset(self):
        qs = super().get_queryset()
        category = self.request.query_params.get('category')
        if category:
            qs = qs.filter(category__name__iexact=category)
        return qs

from rest_framework.generics import RetrieveUpdateAPIView

class ProductDetailView(RetrieveUpdateAPIView):
    """
    GET /api/products/<id>/ -> get product
    PATCH /api/products/<id>/ -> update product (supports image upload)
    """
    permission_classes = [IsAdminUser]
    serializer_class = ProductSerializer
    queryset = Product.objects.all()
    parser_classes = [MultiPartParser, FormParser, JSONParser]


# ---------- 1) Upload Excel -> Update stock/price ----------
from .utils import process_stock_excel

# ---------- 1) Upload Excel -> Update stock/price ----------
class StockExcelUploadView(APIView):
    """
    POST multipart/form-data:
      - file: REPORT.xlsx (specific format)
    """
    permission_classes = [IsAdminUser]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        f = request.FILES.get("file")
        if not f:
            return Response({"detail": "Upload 'file' (xlsx)."}, status=400)

        # process_stock_excel expects a file path or file-like object
        # It uses openpyxl.load_workbook which accepts file-like objects (BytesIO)
        # request.FILES['file'] is likely an InMemoryUploadedFile or TemporaryUploadedFile
        
        try:
             results = process_stock_excel(f)
        except Exception as e:
             return Response({"detail": f"Error processing file: {str(e)}"}, status=400)

        return Response({
            "ok": True,
            "results": results
        })


# ---------- 2) Download Excel of current stock ----------
class StockExcelDownloadView(APIView):
    permission_classes = [IsAdminUser]

    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "Stock"
        ws.append(["ID", "Name", "Category", "Price", "Stock"])
        qs = Product.objects.select_related("category").order_by("name")
        for p in qs:
            ws.append([
                p.id,
                p.name,
                p.category.name if p.category_id else "",
                str(p.price),   # keep as string to avoid locale issues
                p.stock,
            ])

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f'stock_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        # FileResponse streams efficiently
        resp = FileResponse(
            buf,
            as_attachment=True,
            filename=filename,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        return resp


# ---------- 3) Bulk edit stock/price ----------
class ProductBulkUpdateView(APIView):
    """
    PATCH body:
    {
      "items": [
        {"id": 1, "stock": 10},
        {"id": 2, "price": 99.90},
        {"id": 3, "stock": 8, "price": 15.50}
      ]
    }
    """
    permission_classes = [IsAdminUser]

    def patch(self, request):
        ser = ProductBulkUpdateSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        items = ser.validated_data["items"]

        updated, errors = 0, []
        with transaction.atomic():
            for it in items:
                pid = it["id"]
                try:
                    obj = Product.objects.select_for_update().get(pk=pid)
                    fields = []
                    if "stock" in it:
                        obj.stock = int(it["stock"])
                        fields.append("stock")
                    if "price" in it:
                        obj.price = it["price"]
                        fields.append("price")
                    if fields:
                        obj.save(update_fields=fields)
                        updated += 1
                except Product.DoesNotExist:
                    errors.append(f"Product {pid} not found")
                except Exception as e:
                    errors.append(f"Product {pid}: {e}")

        return Response({"ok": True, "updated": updated, "errors": errors})


# ---------- 4) Daily sales report (JSON or ?format=xlsx) ----------
# Uses Order / OrderItem from your orders app.
from orders.models import Order, OrderItem, OrderStatus, OrderSource

# class DailySalesReportView(APIView):
#     """
#     GET params:
#       - date=YYYY-MM-DD (default: today)
#       - format=xlsx -> download Excel, else JSON
#     """
#     permission_classes = [IsAdminUser]
#
#     def get(self, request):
#         tz = get_current_timezone()
#         date_str = request.query_params.get("date")
#         if date_str:
#             y, m, d = map(int, date_str.split("-"))
#             day = date(y, m, d)
#         else:
#             day = date.today()
#
#         start = make_aware(datetime.combine(day, datetime.min.time()), tz)
#         end = start + timedelta(days=1)
#
#         paid_like = [OrderStatus.PAID, OrderStatus.SHIPPED, OrderStatus.COMPLETED]
#
#         orders = (
#             Order.objects
#             .filter(created_at__gte=start, created_at__lt=end, status__in=paid_like)
#             .select_related("user")
#         )
#
#         # Totals
#         orders_count = orders.count()
#         revenue_total = orders.aggregate(s=Sum("total_amount"))["s"] or 0
#
#         # Split by source
#         src = (
#             orders.values("source")
#             .annotate(orders=Sum(1), revenue=Sum("total_amount"))
#             .order_by()
#         )
#         by_source = {row["source"]: {"orders": row["orders"], "revenue": float(row["revenue"] or 0)} for row in src}
#
#         # Items aggregation
#         items = (
#             OrderItem.objects
#             .filter(order__in=orders)
#             .select_related("product__category")
#             .values("product_id", "product__name", "product__category__name")
#             .annotate(qty=Sum("quantity"), amount=Sum("line_total"))
#             .order_by("-qty")
#         )
#         by_product = [
#             {
#                 "product_id": r["product_id"],
#                 "name": r["product__name"],
#                 "category": r["product__category__name"],
#                 "quantity": int(r["qty"] or 0),
#                 "amount": float(r["amount"] or 0),
#             }
#             for r in items
#         ]
#         units_sold = sum(x["quantity"] for x in by_product)
#
#         if request.query_params.get("format") == "xlsx":
#             # Build Excel
#             wb = Workbook()
#             ws1 = wb.active
#             ws1.title = "Summary"
#             ws1.append(["Date", str(day)])
#             ws1.append(["Orders", orders_count])
#             ws1.append(["Units sold", units_sold])
#             ws1.append(["Revenue", float(revenue_total)])
#             ws1.append([])
#             ws1.append(["Source", "Orders", "Revenue"])
#             for s in (OrderSource.ONLINE, OrderSource.POS):
#                 vals = by_source.get(s, {"orders": 0, "revenue": 0.0})
#                 ws1.append([s, vals["orders"], vals["revenue"]])
#
#             ws2 = wb.create_sheet("By Product")
#             ws2.append(["product_id", "name", "category", "quantity", "amount"])
#             for r in by_product:
#                 ws2.append([r["product_id"], r["name"], r["category"], r["quantity"], r["amount"]])
#
#             buf = BytesIO()
#             wb.save(buf); buf.seek(0)
#             resp = HttpResponse(
#                 buf.getvalue(),
#                 content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#             )
#             resp["Content-Disposition"] = f'attachment; filename="sales_{day.isoformat()}.xlsx"'
#             return resp
#
#         return Response({
#             "date": day.isoformat(),
#             "orders": orders_count,
#             "units_sold": units_sold,
#             "revenue": float(revenue_total),
#             "by_source": by_source,
#             "by_product": by_product[:500],  # cap
#         })
class DailySalesReportView(APIView):
    """
    GET params:
      - date=YYYY-MM-DD (default: today)
      - format=xlsx -> download Excel, else JSON
    """
    permission_classes = [IsAdminUser]

    def get(self, request):
        tz = get_current_timezone()
        date_str = request.query_params.get("date")
        if date_str:
            y, m, d = map(int, date_str.split("-"))
            day = date(y, m, d)
        else:
            day = date.today()

        start = make_aware(datetime.combine(day, datetime.min.time()), tz)
        end = start + timedelta(days=1)

        paid_like = [OrderStatus.PAID, OrderStatus.SHIPPED, OrderStatus.COMPLETED]

        orders = (
            Order.objects
            .filter(created_at__gte=start, created_at__lt=end, status__in=paid_like)
            .select_related("user")
        )

        # Totals
        orders_count = orders.count()
        revenue_total = orders.aggregate(s=Sum("total_amount"))["s"] or 0

        # Split by source
        src = (
            orders.values("source")
            .annotate(orders=Count("id"), revenue=Sum("total_amount"))
            .order_by()
        )
        by_source = {
            row["source"]: {
                "orders": row["orders"],
                "revenue": float(row["revenue"] or 0)
            } for row in src
        }

        # Split by cashier (POS vs ONLINE)
        invoices = (
            BillingInvoice.objects
            .filter(created_at__gte=start, created_at__lt=end, status=BillingInvoice.STATUS_PAID)
            .select_related("cashier", "order")
        )

        by_cashier = []
        for inv in invoices:
            if inv.mode == BillingInvoice.MODE_MANUAL:  # POS / Offline
                cashier_name = inv.cashier.get_username() if inv.cashier else "Unknown"
            else:  # Online order (no cashier, fallback to order.user if needed)
                cashier_name = inv.order.user.get_username() if inv.order and inv.order.user else "N/A"

            by_cashier.append({
                "invoice_id": inv.id,
                "cashier": cashier_name,
                "mode": inv.mode,
                "amount": float(inv.total or 0),
            })

        # Items aggregation
        items = (
            OrderItem.objects
            .filter(order__in=orders)
            .select_related("product__category")
            .values("product_id", "product__name", "product__category__name")
            .annotate(qty=Sum("quantity"), amount=Sum("line_total"))
            .order_by("-qty")
        )
        by_product = [
            {
                "product_id": r["product_id"],
                "name": r["product__name"],
                "category": r["product__category__name"],
                "quantity": int(r["qty"] or 0),
                "amount": float(r["amount"] or 0),
            }
            for r in items
        ]
        units_sold = sum(x["quantity"] for x in by_product)

        # Excel export
        if request.query_params.get("format") == "xlsx":
            wb = Workbook()
            ws1 = wb.active
            ws1.title = "Summary"
            ws1.append(["Date", str(day)])
            ws1.append(["Orders", orders_count])
            ws1.append(["Units sold", units_sold])
            ws1.append(["Revenue", float(revenue_total)])
            ws1.append([])

            ws1.append(["Source", "Orders", "Revenue"])
            for s in (OrderSource.ONLINE, OrderSource.POS):
                vals = by_source.get(s, {"orders": 0, "revenue": 0.0})
                ws1.append([s, vals["orders"], vals["revenue"]])

            ws2 = wb.create_sheet("By Product")
            ws2.append(["product_id", "name", "category", "quantity", "amount"])
            for r in by_product:
                ws2.append([r["product_id"], r["name"], r["category"], r["quantity"], r["amount"]])

            ws3 = wb.create_sheet("By Cashier")
            ws3.append(["invoice_id", "cashier", "mode", "amount"])
            for r in by_cashier:
                ws3.append([r["invoice_id"], r["cashier"], r["mode"], r["amount"]])

            buf = BytesIO()
            wb.save(buf); buf.seek(0)
            resp = HttpResponse(
                buf.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            resp["Content-Disposition"] = f'attachment; filename="sales_{day.isoformat()}.xlsx"'
            return resp

        # JSON response
        return Response({
            "date": day.isoformat(),
            "orders": orders_count,
            "units_sold": units_sold,
            "revenue": float(revenue_total),
            "by_source": by_source,
            "by_cashier": by_cashier,
            "by_product": by_product[:500],  # cap
        })